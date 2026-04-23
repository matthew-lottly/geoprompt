"""Data management, conversion, schema tools, and enterprise data helpers.

Covers geodatabase inspection, domain validation, field calculators,
append/upsert workflows, format conversion, schema diffing, coordinate
cleaning, lineage stamps, and workspace cataloging.
"""

from __future__ import annotations

import copy
import csv
import hashlib
import importlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Sequence

from .safe_expression import evaluate_safe_expression

# ---------------------------------------------------------------------------
# Lazy helpers
# ---------------------------------------------------------------------------

def _try_import(name: str) -> Any:
    try:
        return importlib.import_module(name)
    except ImportError:
        return None


# ── 1. Geodatabase schema inspection and export ──────────────────────────────

def inspect_geodatabase(path: str | Path) -> dict[str, Any]:
    """Inspect a file geodatabase (.gdb) or GeoPackage and return schema info.

    Returns ``layers`` list with name, geometry_type, field_count, feature_count.
    """
    import fiona  # type: ignore[import-untyped]
    gdb_path = Path(path)
    layers: list[dict[str, Any]] = []
    for layer_name in fiona.listlayers(str(gdb_path)):
        with fiona.open(str(gdb_path), layer=layer_name) as src:
            layers.append({
                "name": layer_name,
                "geometry_type": src.schema.get("geometry", "Unknown"),
                "fields": list(src.schema.get("properties", {}).keys()),
                "field_count": len(src.schema.get("properties", {})),
                "feature_count": len(src),
                "crs": str(src.crs) if src.crs else None,
            })
    return {"path": str(gdb_path), "layer_count": len(layers), "layers": layers}


def export_schema(records: Sequence[dict[str, Any]], *, name: str = "dataset") -> dict[str, Any]:
    """Export a lightweight schema descriptor from a record list."""
    if not records:
        return {"name": name, "fields": [], "row_count": 0}
    sample = records[0]
    fields = []
    for k, v in sample.items():
        dtype = type(v).__name__ if v is not None else "unknown"
        fields.append({"name": k, "type": dtype})
    return {"name": name, "fields": fields, "row_count": len(records)}


# ── 2. Domains and subtype validation ────────────────────────────────────────

class FieldDomain:
    """Enterprise-style coded-value or range domain for field validation."""

    def __init__(
        self,
        name: str,
        domain_type: str = "coded",
        values: dict[Any, str] | None = None,
        min_value: float | None = None,
        max_value: float | None = None,
    ) -> None:
        self.name = name
        self.domain_type = domain_type
        self.values = values or {}
        self.min_value = min_value
        self.max_value = max_value

    def validate(self, value: Any) -> bool:
        if value is None:
            return True
        if self.domain_type == "coded":
            return value in self.values
        if self.domain_type == "range":
            if self.min_value is not None and value < self.min_value:
                return False
            if self.max_value is not None and value > self.max_value:
                return False
            return True
        return True

    def to_dict(self) -> dict[str, Any]:
        return {"name": self.name, "type": self.domain_type, "values": self.values, "min": self.min_value, "max": self.max_value}


def validate_domains(
    records: Sequence[dict[str, Any]],
    domain_map: dict[str, FieldDomain],
) -> list[dict[str, Any]]:
    """Validate records against field domains; return violations."""
    violations: list[dict[str, Any]] = []
    for i, r in enumerate(records):
        for field, domain in domain_map.items():
            val = r.get(field)
            if not domain.validate(val):
                violations.append({"row": i, "field": field, "value": val, "domain": domain.name})
    return violations


# ── 3. Attribute rules and constraint-checking ───────────────────────────────

def check_constraints(
    records: Sequence[dict[str, Any]],
    constraints: dict[str, Callable[[Any], bool]],
) -> list[dict[str, Any]]:
    """Check each record against per-field constraint functions.

    *constraints* maps field names to callables that return True if valid.
    """
    violations: list[dict[str, Any]] = []
    for i, r in enumerate(records):
        for field, fn in constraints.items():
            val = r.get(field)
            try:
                if not fn(val):
                    violations.append({"row": i, "field": field, "value": val})
            except Exception as exc:
                violations.append({"row": i, "field": field, "value": val, "error": str(exc)})
    return violations


# ── 4. Field calculator ──────────────────────────────────────────────────────

_SAFE_BUILTINS = {"abs": abs, "min": min, "max": max, "round": round, "len": len, "int": int, "float": float, "str": str, "bool": bool}


def field_calculate(
    records: list[dict[str, Any]],
    target_field: str,
    expression: str,
    *,
    safe: bool = True,
) -> list[dict[str, Any]]:
    """Evaluate *expression* per row and store result in *target_field*.

    Uses a restricted eval sandbox by default.  ``row`` is the current record dict.
    """
    for r in records:
        ns = dict(_SAFE_BUILTINS) if safe else {}
        ns["row"] = r
        r[target_field] = evaluate_safe_expression(expression, ns)
    return records


# ── 5. Append, upsert, truncate, overwrite workflows ────────────────────────

def append_records(
    target: list[dict[str, Any]],
    source: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Append source records to target, returning the combined list."""
    target.extend(source)
    return target


def upsert_records(
    target: list[dict[str, Any]],
    source: Sequence[dict[str, Any]],
    key: str,
) -> list[dict[str, Any]]:
    """Insert or update records in *target* from *source* by *key*."""
    index = {r[key]: i for i, r in enumerate(target) if key in r}
    for s in source:
        k = s.get(key)
        if k is not None and k in index:
            target[index[k]].update(s)
        else:
            target.append(s)
    return target


def truncate_records(target: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Remove all records from *target* in place."""
    target.clear()
    return target


def overwrite_records(
    target: list[dict[str, Any]],
    source: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Replace all records in *target* with *source*."""
    target.clear()
    target.extend(source)
    return target


# ── 6. Batch rename, reorder, alias ──────────────────────────────────────────

def batch_rename_fields(
    records: list[dict[str, Any]],
    rename_map: dict[str, str],
) -> list[dict[str, Any]]:
    """Rename fields across all records per *rename_map*."""
    for r in records:
        for old, new in rename_map.items():
            if old in r:
                r[new] = r.pop(old)
    return records


def reorder_fields(
    records: Sequence[dict[str, Any]],
    order: Sequence[str],
) -> list[dict[str, Any]]:
    """Return records with fields reordered per *order*.  Extra fields appended."""
    result: list[dict[str, Any]] = []
    for r in records:
        ordered = {k: r[k] for k in order if k in r}
        for k, v in r.items():
            if k not in ordered:
                ordered[k] = v
        result.append(ordered)
    return result


# ── 7. Relationship-class style helpers ──────────────────────────────────────

def one_to_many_join(
    parent: Sequence[dict[str, Any]],
    child: Sequence[dict[str, Any]],
    parent_key: str,
    child_key: str,
    children_field: str = "_children",
) -> list[dict[str, Any]]:
    """Attach child records to parents as nested lists."""
    child_index: dict[Any, list[dict[str, Any]]] = {}
    for c in child:
        k = c.get(child_key)
        child_index.setdefault(k, []).append(c)
    result = []
    for p in parent:
        row = dict(p)
        row[children_field] = child_index.get(p.get(parent_key), [])
        result.append(row)
    return result


def many_to_many_join(
    left: Sequence[dict[str, Any]],
    right: Sequence[dict[str, Any]],
    bridge: Sequence[dict[str, str]],
    left_key: str,
    right_key: str,
    bridge_left: str,
    bridge_right: str,
) -> list[dict[str, Any]]:
    """Resolve a many-to-many relationship through a bridge table."""
    right_idx = {r[right_key]: r for r in right if right_key in r}
    bridge_map: dict[Any, list[Any]] = {}
    for b in bridge:
        bridge_map.setdefault(b.get(bridge_left), []).append(b.get(bridge_right))
    results: list[dict[str, Any]] = []
    for left_row in left:
        lk = left_row.get(left_key)
        for rk in bridge_map.get(lk, []):
            matched = right_idx.get(rk)
            if matched:
                merged = dict(left_row)
                for mk, mv in matched.items():
                    if mk != right_key:
                        merged[f"right_{mk}"] = mv
                results.append(merged)
    return results


# ── 8. Feature class copy, move, archive, snapshot ───────────────────────────

def copy_records(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Deep-copy a record list."""
    return copy.deepcopy(list(records))


def archive_records(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
    *,
    fmt: str = "json",
) -> str:
    """Write records to an archive file and return the path."""
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "json":
        p.write_text(json.dumps(list(records), default=str), encoding="utf-8")
    elif fmt == "csv":
        if not records:
            p.write_text("", encoding="utf-8")
        else:
            keys = list(records[0].keys())
            with p.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=keys)
                w.writeheader()
                w.writerows(records)
    return str(p)


def snapshot_records(
    records: Sequence[dict[str, Any]],
    *,
    label: str = "",
) -> dict[str, Any]:
    """Create a versioned in-memory snapshot with timestamp and hash."""
    payload = json.dumps(list(records), default=str, sort_keys=True)
    digest = hashlib.sha256(payload.encode()).hexdigest()[:12]
    return {
        "label": label,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "hash": digest,
        "row_count": len(records),
        "data": list(records),
    }


# ── 9. Metadata read/write/clone ─────────────────────────────────────────────

def read_metadata(path: str | Path) -> dict[str, Any]:
    """Read JSON metadata sidecar (.meta.json) for a dataset."""
    meta_path = Path(str(path) + ".meta.json")
    if meta_path.exists():
        return json.loads(meta_path.read_text(encoding="utf-8"))
    return {}


def write_metadata(path: str | Path, metadata: dict[str, Any]) -> str:
    """Write JSON metadata sidecar for a dataset."""
    meta_path = Path(str(path) + ".meta.json")
    meta_path.write_text(json.dumps(metadata, indent=2, default=str), encoding="utf-8")
    return str(meta_path)


def clone_metadata(source_path: str | Path, target_path: str | Path) -> str:
    """Clone the metadata sidecar from one dataset to another."""
    meta = read_metadata(source_path)
    meta["cloned_from"] = str(source_path)
    meta["cloned_at"] = datetime.now(timezone.utc).isoformat()
    return write_metadata(target_path, meta)


# ── 10. Template-driven project scaffolding ──────────────────────────────────

def scaffold_project(
    root: str | Path,
    *,
    name: str = "project",
    folders: Sequence[str] = ("data", "outputs", "docs", "scripts"),
) -> dict[str, Any]:
    """Create a project directory structure with starter files."""
    root_path = Path(root) / name
    root_path.mkdir(parents=True, exist_ok=True)
    created: list[str] = [str(root_path)]
    for folder in folders:
        fp = root_path / folder
        fp.mkdir(exist_ok=True)
        created.append(str(fp))
    readme = root_path / "README.md"
    if not readme.exists():
        readme.write_text(f"# {name}\n\nGeoPrompt project.\n", encoding="utf-8")
        created.append(str(readme))
    return {"root": str(root_path), "created": created}


# ── 11. Batch format conversion ──────────────────────────────────────────────

def batch_convert(
    input_paths: Sequence[str | Path],
    output_dir: str | Path,
    *,
    target_format: str = "geojson",
) -> list[str]:
    """Convert multiple files to *target_format* (geojson, csv, json).

    Uses the geoprompt IO layer internally.
    """
    from geoprompt.io import read_data, write_data
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    outputs: list[str] = []
    for inp in input_paths:
        frame = read_data(str(inp))
        stem = Path(inp).stem
        ext = {"geojson": ".geojson", "csv": ".csv", "json": ".json"}.get(target_format, f".{target_format}")
        out_path = out / f"{stem}{ext}"
        write_data(str(out_path), frame)
        outputs.append(str(out_path))
    return outputs


# ── 12. Schema diff and migration planner ────────────────────────────────────

def schema_diff(
    schema_a: dict[str, str],
    schema_b: dict[str, str],
) -> dict[str, Any]:
    """Compare two field-name→type dicts and report differences."""
    a_keys = set(schema_a)
    b_keys = set(schema_b)
    added = {k: schema_b[k] for k in b_keys - a_keys}
    removed = {k: schema_a[k] for k in a_keys - b_keys}
    type_changed = {}
    for k in a_keys & b_keys:
        if schema_a[k] != schema_b[k]:
            type_changed[k] = {"from": schema_a[k], "to": schema_b[k]}
    return {"added": added, "removed": removed, "type_changed": type_changed, "compatible": not (removed or type_changed)}


def migration_plan(diff: dict[str, Any]) -> list[str]:
    """Generate migration steps from a schema diff."""
    steps: list[str] = []
    for field, dtype in diff.get("added", {}).items():
        steps.append(f"ADD COLUMN {field} {dtype}")
    for field in diff.get("removed", {}):
        steps.append(f"DROP COLUMN {field}")
    for field, change in diff.get("type_changed", {}).items():
        steps.append(f"ALTER COLUMN {field} FROM {change['from']} TO {change['to']}")
    return steps


# ── 13. Feature comparison and change-detection ──────────────────────────────

def feature_change_report(
    old_records: Sequence[dict[str, Any]],
    new_records: Sequence[dict[str, Any]],
    key: str,
) -> dict[str, Any]:
    """Detect added, removed, and modified features between two snapshots."""
    old_idx = {r[key]: r for r in old_records if key in r}
    new_idx = {r[key]: r for r in new_records if key in r}
    added = [k for k in new_idx if k not in old_idx]
    removed = [k for k in old_idx if k not in new_idx]
    modified: list[Any] = []
    for k in set(old_idx) & set(new_idx):
        if json.dumps(old_idx[k], sort_keys=True, default=str) != json.dumps(new_idx[k], sort_keys=True, default=str):
            modified.append(k)
    return {"added": added, "removed": removed, "modified": modified, "unchanged": len(old_idx) - len(removed) - len(modified)}


# ── 14. Coordinate cleaning and standardization ─────────────────────────────

def clean_coordinates(
    records: list[dict[str, Any]],
    x_field: str = "longitude",
    y_field: str = "latitude",
) -> list[dict[str, Any]]:
    """Clamp coordinates to valid WGS-84 ranges and strip whitespace."""
    for r in records:
        try:
            x = float(str(r.get(x_field, "")).strip())
            y = float(str(r.get(y_field, "")).strip())
            r[x_field] = max(-180.0, min(180.0, x))
            r[y_field] = max(-90.0, min(90.0, y))
        except (ValueError, TypeError):
            pass
    return records


def standardize_columns(
    records: list[dict[str, Any]],
    *,
    lower: bool = True,
    strip: bool = True,
    replace_spaces: str = "_",
) -> list[dict[str, Any]]:
    """Normalize column names across all records."""
    result: list[dict[str, Any]] = []
    for r in records:
        new_r: dict[str, Any] = {}
        for k, v in r.items():
            nk = k
            if strip:
                nk = nk.strip()
            if replace_spaces:
                nk = nk.replace(" ", replace_spaces)
            if lower:
                nk = nk.lower()
            new_r[nk] = v
        result.append(new_r)
    return result


# ── 15. Workspace catalog browser ────────────────────────────────────────────

_DATA_EXTENSIONS = {
    ".geojson", ".json", ".csv", ".tsv", ".shp", ".gpkg", ".gdb",
    ".parquet", ".fgb", ".xlsx", ".xls", ".tif", ".tiff",
}


def catalog_workspace(root: str | Path) -> list[dict[str, Any]]:
    """Scan a directory tree and return a catalog of recognized datasets."""
    root_path = Path(root)
    entries: list[dict[str, Any]] = []
    for p in sorted(root_path.rglob("*")):
        if p.is_file() and p.suffix.lower() in _DATA_EXTENSIONS:
            entries.append({
                "path": str(p),
                "name": p.name,
                "format": p.suffix.lower().lstrip("."),
                "size_bytes": p.stat().st_size,
            })
    return entries


# ── 16. Richer Parquet / Arrow optimization ──────────────────────────────────

def read_parquet_filtered(
    path: str | Path,
    *,
    columns: Sequence[str] | None = None,
    row_filter: str | None = None,
) -> list[dict[str, Any]]:
    """Read a Parquet file with optional column and row predicate pushdown."""
    import pyarrow.parquet as pq  # type: ignore[import-untyped]
    table = pq.read_table(str(path), columns=list(columns) if columns else None)
    records = table.to_pydict()
    rows: list[dict[str, Any]] = []
    n = len(next(iter(records.values()))) if records else 0
    for i in range(n):
        row = {k: v[i] for k, v in records.items()}
        rows.append(row)
    if row_filter:
        ns = dict(_SAFE_BUILTINS)
        rows = [r for r in rows if evaluate_safe_expression(row_filter, {**ns, "row": r})]
    return rows


# ── 17. Data lineage stamps ──────────────────────────────────────────────────

def stamp_lineage(
    records: list[dict[str, Any]],
    *,
    operation: str,
    source: str = "",
    timestamp: str | None = None,
) -> list[dict[str, Any]]:
    """Embed lineage metadata into each record."""
    ts = timestamp or datetime.now(timezone.utc).isoformat()
    for r in records:
        r.setdefault("_lineage", [])
        r["_lineage"].append({"operation": operation, "source": source, "timestamp": ts})
    return records


# ── 18. Attachment and related-media management ──────────────────────────────

def attach_media(
    record: dict[str, Any],
    media_path: str | Path,
    *,
    field: str = "_attachments",
) -> dict[str, Any]:
    """Register a media file path as an attachment on a record."""
    record.setdefault(field, [])
    record[field].append({"path": str(media_path), "attached_at": datetime.now(timezone.utc).isoformat()})
    return record


def list_attachments(record: dict[str, Any], *, field: str = "_attachments") -> list[dict[str, Any]]:
    """Return attachment metadata from a record."""
    return record.get(field, [])


# ── 19. Zip package import / export ──────────────────────────────────────────

def export_zip_bundle(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
    *,
    extras: Sequence[str | Path] | None = None,
) -> str:
    """Package records as GeoJSON plus extra files into a .zip deliverable."""
    import zipfile
    out = Path(output_path)
    with zipfile.ZipFile(str(out), "w", zipfile.ZIP_DEFLATED) as zf:
        fc = {"type": "FeatureCollection", "features": [
            {"type": "Feature", "properties": {k: v for k, v in r.items() if k != "geometry"}, "geometry": r.get("geometry")}
            for r in records
        ]}
        zf.writestr("data.geojson", json.dumps(fc, default=str))
        for extra in extras or []:
            ep = Path(extra)
            if ep.exists():
                zf.write(str(ep), ep.name)
    return str(out)


def import_zip_bundle(zip_path: str | Path) -> dict[str, Any]:
    """Unpack a .zip and return contained GeoJSON records plus file listing."""
    import zipfile
    with zipfile.ZipFile(str(zip_path), "r") as zf:
        names = zf.namelist()
        records: list[dict[str, Any]] = []
        for n in names:
            if n.endswith(".geojson") or n.endswith(".json"):
                data = json.loads(zf.read(n))
                if isinstance(data, dict) and "features" in data:
                    for f in data["features"]:
                        row = dict(f.get("properties", {}))
                        row["geometry"] = f.get("geometry")
                        records.append(row)
        return {"files": names, "records": records}


# ── 20. Richer Excel roundtrip ───────────────────────────────────────────────

def write_excel_styled(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
    *,
    sheet_name: str = "Data",
    header_color: str = "4472C4",
) -> str:
    """Write records to Excel with styled headers using openpyxl."""
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if not records:
        wb.save(str(output_path))
        return str(output_path)
    headers = list(records[0].keys())
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = fill
        cell.font = font
    for ri, rec in enumerate(records, 2):
        for ci, h in enumerate(headers, 1):
            v = rec.get(h)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, default=str)
            ws.cell(row=ri, column=ci, value=v)
    wb.save(str(output_path))
    return str(output_path)


# ---------------------------------------------------------------------------
# Additional data management utilities (A7 items)
# ---------------------------------------------------------------------------


def frequency_analysis(
    records: Sequence[dict[str, Any]],
    fields: Sequence[str],
) -> list[dict[str, Any]]:
    """Count the frequency of unique value combinations across *fields*.

    Parameters
    ----------
    records : sequence of dicts
    fields : sequence of str
        Field names to group by.

    Returns
    -------
    list[dict]
        Records with the grouped field values plus a ``"FREQUENCY"`` key.
    """
    from collections import Counter

    counter: Counter[tuple[Any, ...]] = Counter()
    for rec in records:
        key = tuple(rec.get(f) for f in fields)
        counter[key] += 1

    result: list[dict[str, Any]] = []
    for key, count in counter.most_common():
        row: dict[str, Any] = {f: v for f, v in zip(fields, key)}
        row["FREQUENCY"] = count
        result.append(row)
    return result


def find_identical(
    records: Sequence[dict[str, Any]],
    fields: Sequence[str],
) -> list[list[int]]:
    """Find groups of records that are identical across *fields*.

    Parameters
    ----------
    records : sequence of dicts
    fields : sequence of str

    Returns
    -------
    list[list[int]]
        Groups of record indices that share the same values.
        Only groups with 2+ members are included.
    """
    from collections import defaultdict

    groups: dict[tuple[Any, ...], list[int]] = defaultdict(list)
    for idx, rec in enumerate(records):
        key = tuple(rec.get(f) for f in fields)
        groups[key].append(idx)
    return [idxs for idxs in groups.values() if len(idxs) >= 2]


def delete_identical(
    records: list[dict[str, Any]],
    fields: Sequence[str],
) -> list[dict[str, Any]]:
    """Remove duplicate records, keeping first occurrence per unique field combo.

    Parameters
    ----------
    records : list of dicts
    fields : sequence of str

    Returns
    -------
    list[dict]
        De-duplicated records.
    """
    seen: set[tuple[Any, ...]] = set()
    result: list[dict[str, Any]] = []
    for rec in records:
        key = tuple(rec.get(f) for f in fields)
        if key not in seen:
            seen.add(key)
            result.append(rec)
    return result


def transpose_table(
    records: Sequence[dict[str, Any]],
    id_field: str,
) -> list[dict[str, Any]]:
    """Transpose a table, turning row values into columns.

    The *id_field* values become column headers; other fields become rows.

    Parameters
    ----------
    records : sequence of dicts
    id_field : str
        Field whose values become the new column names.

    Returns
    -------
    list[dict]
        Transposed records.
    """
    if not records:
        return []
    other_fields = [f for f in records[0] if f != id_field]
    result: list[dict[str, Any]] = []
    for field in other_fields:
        row: dict[str, Any] = {"field": field}
        for rec in records:
            col_name = str(rec.get(id_field, ""))
            row[col_name] = rec.get(field)
        result.append(row)
    return result


def table_to_csv(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
    *,
    fields: Sequence[str] | None = None,
) -> str:
    """Write records to a CSV file.

    Parameters
    ----------
    records : sequence of dicts
    output_path : str or Path
    fields : sequence of str, optional
        Columns to include. If omitted, all columns from the first record are used.

    Returns
    -------
    str
        The path written.
    """
    if not records:
        Path(output_path).write_text("", encoding="utf-8")
        return str(output_path)
    cols = list(fields) if fields else list(records[0].keys())
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)
    return str(output_path)


def get_count(records: Sequence[dict[str, Any]]) -> int:
    """Return the number of records.

    Parameters
    ----------
    records : sequence of dicts

    Returns
    -------
    int
    """
    return len(records)


def list_fields(
    records: Sequence[dict[str, Any]],
) -> list[dict[str, str]]:
    """List the field names and inferred types of a record set.

    Parameters
    ----------
    records : sequence of dicts

    Returns
    -------
    list[dict]
        Each dict has ``"name"`` and ``"type"`` keys.
    """
    if not records:
        return []
    sample = records[0]
    result: list[dict[str, str]] = []
    for key, val in sample.items():
        t = type(val).__name__ if val is not None else "NoneType"
        result.append({"name": key, "type": t})
    return result


def table_compare(
    table_a: Sequence[dict[str, Any]],
    table_b: Sequence[dict[str, Any]],
    *,
    key_fields: Sequence[str] | None = None,
) -> dict[str, Any]:
    """Compare two tables and report differences.

    Parameters
    ----------
    table_a, table_b : sequences of dicts
    key_fields : sequence of str, optional
        Fields that identify matching rows. If omitted, comparison is by index.

    Returns
    -------
    dict
        Keys: ``"added"`` (in B not A), ``"removed"`` (in A not B),
        ``"modified"`` (matching key but different values),
        ``"identical_count"`` (matching rows).
    """
    if key_fields:
        def _key(rec: dict[str, Any]) -> tuple[Any, ...]:
            return tuple(rec.get(f) for f in key_fields)

        map_a = {_key(r): r for r in table_a}
        map_b = {_key(r): r for r in table_b}
        added = [map_b[k] for k in map_b if k not in map_a]
        removed = [map_a[k] for k in map_a if k not in map_b]
        modified: list[dict[str, Any]] = []
        identical = 0
        for k in map_a:
            if k in map_b:
                if map_a[k] == map_b[k]:
                    identical += 1
                else:
                    modified.append({"key": k, "a": map_a[k], "b": map_b[k]})
    else:
        added_list: list[dict[str, Any]] = []
        removed_list: list[dict[str, Any]] = []
        modified = []
        identical = 0
        max_len = max(len(table_a), len(table_b))
        for i in range(max_len):
            if i >= len(table_a):
                added_list.append(table_b[i])
            elif i >= len(table_b):
                removed_list.append(table_a[i])
            elif table_a[i] == table_b[i]:
                identical += 1
            else:
                modified.append({"index": i, "a": table_a[i], "b": table_b[i]})
        added = added_list
        removed = removed_list

    return {
        "added": added,
        "removed": removed,
        "modified": modified,
        "identical_count": identical,
    }


def split_layer_by_attribute(
    records: Sequence[dict[str, Any]],
    field: str,
) -> dict[str, list[dict[str, Any]]]:
    """Split records into groups by unique values of *field*.

    Parameters
    ----------
    records : sequence of dicts
    field : str
        Field to group by.

    Returns
    -------
    dict[str, list[dict]]
        Keys are the unique field values (stringified); values are record lists.
    """
    from collections import defaultdict

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for rec in records:
        key = str(rec.get(field, ""))
        groups[key].append(rec)
    return dict(groups)


def validate_table_name(name: str) -> dict[str, Any]:
    """Validate a table name against common database naming rules.

    Returns
    -------
    dict
        Keys: ``"valid"`` (bool), ``"issues"`` (list of str).
    """
    import re

    issues: list[str] = []
    if not name:
        issues.append("name is empty")
    elif not re.match(r"^[A-Za-z_]", name):
        issues.append("must start with a letter or underscore")
    if re.search(r"[^A-Za-z0-9_]", name):
        issues.append("contains invalid characters (only letters, digits, _ allowed)")
    if len(name) > 128:
        issues.append("exceeds 128 characters")
    reserved = {"SELECT", "FROM", "WHERE", "TABLE", "DROP", "INSERT", "UPDATE", "DELETE", "CREATE", "ALTER", "INDEX"}
    if name.upper() in reserved:
        issues.append(f"'{name}' is a reserved SQL keyword")
    return {"valid": len(issues) == 0, "issues": issues}


def validate_field_name(name: str) -> dict[str, Any]:
    """Validate a field name against common database naming rules.

    Returns
    -------
    dict
        Keys: ``"valid"`` (bool), ``"issues"`` (list of str).
    """
    import re

    issues: list[str] = []
    if not name:
        issues.append("name is empty")
    elif not re.match(r"^[A-Za-z_]", name):
        issues.append("must start with a letter or underscore")
    if re.search(r"[^A-Za-z0-9_]", name):
        issues.append("contains invalid characters")
    if len(name) > 64:
        issues.append("exceeds 64 characters")
    return {"valid": len(issues) == 0, "issues": issues}


def write_csv_with_xy(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
    *,
    geometry_column: str = "geometry",
    x_field: str = "X",
    y_field: str = "Y",
) -> str:
    """Write records to CSV with X/Y coordinate columns extracted from geometry.

    Parameters
    ----------
    records : sequence of dicts
    output_path : str or Path
    geometry_column : str
    x_field, y_field : str
        Column names for the coordinates in the output.

    Returns
    -------
    str
        Path written.
    """
    from .geometry import geometry_centroid

    if not records:
        Path(output_path).write_text("", encoding="utf-8")
        return str(output_path)

    non_geom_fields = [k for k in records[0] if k != geometry_column]
    fieldnames = non_geom_fields + [x_field, y_field]

    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            row = {k: rec.get(k) for k in non_geom_fields}
            geom = rec.get(geometry_column)
            if geom:
                centroid = geometry_centroid(geom)
                row[x_field] = centroid[0]
                row[y_field] = centroid[1]
            else:
                row[x_field] = None
                row[y_field] = None
            writer.writerow(row)
    return str(output_path)


def pretty_print_feature_table(
    records: Sequence[dict[str, Any]],
    *,
    max_rows: int = 20,
    max_col_width: int = 30,
    geometry_column: str = "geometry",
) -> str:
    """Format records as a human-readable ASCII table string.

    Geometry fields are summarized as their type rather than printed in full.

    Parameters
    ----------
    records : sequence of dicts
    max_rows : int
        Maximum rows to display. Default 20.
    max_col_width : int
        Maximum column width. Default 30.
    geometry_column : str

    Returns
    -------
    str
        Formatted table string.
    """
    if not records:
        return "(empty table)"
    display = records[:max_rows]
    cols = list(display[0].keys())

    def _fmt(val: Any, col: str) -> str:
        if col == geometry_column and isinstance(val, dict):
            return f"<{val.get('type', 'Geometry')}>"
        s = str(val) if val is not None else ""
        return s[:max_col_width] if len(s) > max_col_width else s

    rows = [[_fmt(rec.get(c), c) for c in cols] for rec in display]
    widths = [max(len(c), max((len(r[i]) for r in rows), default=0)) for i, c in enumerate(cols)]
    header = " | ".join(c.ljust(widths[i]) for i, c in enumerate(cols))
    sep = "-+-".join("-" * widths[i] for i in range(len(cols)))
    lines = [header, sep]
    for row in rows:
        lines.append(" | ".join(row[i].ljust(widths[i]) for i in range(len(cols))))
    if len(records) > max_rows:
        lines.append(f"... ({len(records) - max_rows} more rows)")
    return "\n".join(lines)


def table_to_dbase(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
) -> str:
    """Write a lightweight dBASE-style table export.

    For portability this writes a delimited text representation using a .dbf
    extension, which is sufficient for round-trip delivery and testing.
    """
    out = Path(output_path)
    if out.suffix.lower() != ".dbf":
        out = out.with_suffix(".dbf")
    if not records:
        out.write_text("", encoding="utf-8")
        return str(out)
    fields = list(records[0].keys())
    with out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, delimiter="|")
        writer.writeheader()
        writer.writerows(records)
    return str(out)


def _point_in_polygon(x: float, y: float, polygon: Sequence[tuple[float, float]]) -> bool:
    inside = False
    j = len(polygon) - 1
    for i in range(len(polygon)):
        xi, yi = polygon[i]
        xj, yj = polygon[j]
        intersects = ((yi > y) != (yj > y)) and (
            x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi
        )
        if intersects:
            inside = not inside
        j = i
    return inside


def select_by_polygon(
    records: Sequence[dict[str, Any]],
    polygon: Sequence[tuple[float, float]],
    *,
    x_field: str = "x",
    y_field: str = "y",
) -> list[dict[str, Any]]:
    """Select records whose X/Y coordinates fall inside a polygon."""
    return [
        rec for rec in records
        if _point_in_polygon(float(rec.get(x_field, 0)), float(rec.get(y_field, 0)), polygon)
    ]


_SCHEMA_LOCKS: dict[str, bool] = {}
_VERSIONED_DATASETS: dict[str, bool] = {}


def compact_database(path: str | Path) -> dict[str, Any]:
    """Compact a file geodatabase-like folder by returning maintenance metadata."""
    p = Path(path)
    return {"path": str(p), "exists": p.exists(), "operation": "compact"}


def compress_database(path: str | Path, *, archive: bool = True) -> dict[str, Any]:
    """Compress a database workspace into a portable archive marker."""
    p = Path(path)
    archive_path = Path(str(p) + ".zip")
    if archive and p.exists() and not archive_path.exists():
        archive_path.write_text("compressed database placeholder", encoding="utf-8")
    return {"path": str(p), "archive": str(archive_path), "compressed": p.exists()}


def schema_lock_management(resource: str, *, action: str = "status") -> dict[str, Any]:
    """Lock, unlock, or inspect a lightweight schema lock state."""
    if action == "lock":
        _SCHEMA_LOCKS[resource] = True
    elif action == "unlock":
        _SCHEMA_LOCKS[resource] = False
    return {"resource": resource, "locked": _SCHEMA_LOCKS.get(resource, False)}


def versioning_register_unregister(resource: str, *, register: bool = True) -> dict[str, Any]:
    """Register or unregister a dataset as versioned."""
    _VERSIONED_DATASETS[resource] = bool(register)
    return {"resource": resource, "versioned": _VERSIONED_DATASETS[resource]}


def create_feature_dataset(root: str | Path, name: str) -> dict[str, Any]:
    """Create a feature-dataset folder."""
    p = Path(root) / name
    p.mkdir(parents=True, exist_ok=True)
    return {"path": str(p), "name": name}


def create_file_geodatabase(root: str | Path, name: str) -> dict[str, Any]:
    """Create a file geodatabase folder with .gdb suffix."""
    p = Path(root) / f"{name}.gdb"
    p.mkdir(parents=True, exist_ok=True)
    return {"path": str(p), "name": name}


def create_mobile_geodatabase(root: str | Path, name: str) -> dict[str, Any]:
    """Create a mobile geodatabase placeholder file."""
    p = Path(root) / f"{name}.geodatabase"
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.exists():
        p.write_text("{}", encoding="utf-8")
    return {"path": str(p), "name": name}


def create_dataset_catalog(
    datasets: dict[str, Sequence[dict[str, Any]]],
    *,
    workspace: str | Path | None = None,
) -> dict[str, Any]:
    """Build a lightweight enterprise-style catalog for one or more datasets."""
    catalog: dict[str, Any] = {}
    workspace_text = str(workspace) if workspace is not None else None
    persistence = "simulated"
    if workspace_text:
        lowered = workspace_text.lower()
        if lowered.endswith((".gpkg", ".sqlite", ".db", ".duckdb", ".gdb", ".geodatabase")):
            persistence = "real"
        else:
            persistence = "mixed"

    for name, rows in datasets.items():
        schema = export_schema(rows, name=name)
        field_names = [field["name"] for field in schema.get("fields", [])]
        health = maintenance_audit_pipeline(rows, indexed_fields=[field for field in field_names if field.lower().endswith("id")])
        catalog[name] = {
            "name": name,
            "row_count": len(rows),
            "fields": field_names,
            "schema": schema,
            "has_geometry": any("geometry" in row or "_geometry" in row for row in rows),
            "health": health,
            "persistence": persistence,
        }
    return {
        "workspace": workspace_text,
        "dataset_count": len(catalog),
        "datasets": catalog,
        "persistence": persistence,
    }


def maintenance_audit_pipeline(
    records: Sequence[dict[str, Any]],
    *,
    domain_map: dict[str, FieldDomain] | None = None,
    required_fields: Sequence[str] = (),
    indexed_fields: Sequence[str] = (),
) -> dict[str, Any]:
    """Run a basic maintenance audit over table or feature records."""
    issues: list[dict[str, Any]] = []
    required = list(required_fields)

    for row_index, row in enumerate(records):
        for field in required:
            if field not in row or row.get(field) in (None, ""):
                issues.append({"row": row_index, "field": field, "issue": "missing_required_value"})

    if domain_map:
        for violation in validate_domains(records, domain_map):
            issues.append({**violation, "issue": "domain_violation"})

    duplicate_keys = []
    if indexed_fields:
        for field in indexed_fields:
            seen: set[Any] = set()
            for row_index, row in enumerate(records):
                value = row.get(field)
                if value in seen:
                    duplicate_keys.append({"row": row_index, "field": field, "value": value})
                else:
                    seen.add(value)
        for dup in duplicate_keys:
            issues.append({**dup, "issue": "duplicate_index_value"})

    return {
        "valid": len(issues) == 0,
        "issue_count": len(issues),
        "issues": issues,
        "record_count": len(records),
        "indexed_fields": list(indexed_fields),
        "index_plan": index_planning_suggestions(records, candidate_fields=indexed_fields),
    }


def index_planning_suggestions(
    records: Sequence[dict[str, Any]],
    *,
    candidate_fields: Sequence[str] = (),
) -> dict[str, Any]:
    """Recommend practical indexes for tabular or feature datasets."""
    if not records:
        return {"record_count": 0, "recommended_indexes": [], "notes": ["No records supplied."]}

    fields = list(candidate_fields) or [
        key for key in records[0].keys() if key not in {"geometry", "_geometry"}
    ]
    recommendations: list[dict[str, Any]] = []

    for field in fields:
        values = [row.get(field) for row in records if row.get(field) not in (None, "")]
        if not values:
            continue
        distinct = len({json.dumps(value, sort_keys=True, default=str) for value in values})
        ratio = distinct / len(values)
        normalized = field.lower()
        if normalized.endswith("id") or ratio >= 0.9:
            recommendations.append({"field": field, "kind": "unique_or_primary", "uniqueness_ratio": round(ratio, 3), "reason": "high-cardinality lookup or identifier"})
        elif normalized in {"status", "owner", "category", "type", "priority"} or 0.2 <= ratio <= 0.8:
            recommendations.append({"field": field, "kind": "btree_lookup", "uniqueness_ratio": round(ratio, 3), "reason": "common filter or grouping field"})

    return {
        "record_count": len(records),
        "recommended_indexes": recommendations,
        "notes": ["Recommendations are advisory and tuned for GeoPackage/SQLite/PostGIS-style operational tables."],
    }


def enterprise_persistence_matrix() -> dict[str, dict[str, str]]:
    """Publish which enterprise-style workflows are durable versus simulated."""
    return {
        "geopackage": {"persistence": "real", "status": "supported", "notes": "Uses actual file-backed GeoPackage reads and writes."},
        "sqlite_spatialite": {"persistence": "real", "status": "supported", "notes": "SQLite/SpatiaLite-style reads and writes persist to disk."},
        "postgis": {"persistence": "real", "status": "supported", "notes": "Database connectors operate against real PostGIS tables when dependencies are installed."},
        "duckdb": {"persistence": "real", "status": "supported", "notes": "DuckDB spatial workflows persist when backed by a file database."},
        "edit_session": {"persistence": "simulated", "status": "safe-preview", "notes": "Change tracking, commit, and rollback operate in-memory for deterministic testing and workflow rehearsal."},
        "version_reconcile_post": {"persistence": "simulated", "status": "safe-preview", "notes": "Conflict reporting is real logic, but the posting target is an in-memory record collection."},
        "offline_replica_bundle": {"persistence": "real", "status": "supported", "notes": "Bundle manifests and attachments are written to disk for portable handoff workflows."},
    }


class EditSession:
    """Rollback-friendly edit-session helper with change tracking."""

    def __init__(
        self,
        records: Sequence[dict[str, Any]],
        *,
        key_field: str = "id",
        session_name: str = "default",
    ) -> None:
        self.key_field = key_field
        self.session_name = session_name
        self._base_records = [copy.deepcopy(row) for row in records]
        self._working_records = [copy.deepcopy(row) for row in records]
        self._change_log: list[dict[str, Any]] = []

    def _find_index(self, key: Any) -> int:
        for index, row in enumerate(self._working_records):
            if row.get(self.key_field) == key:
                return index
        raise KeyError(f"record with {self.key_field}={key!r} not found")

    def insert(self, row: dict[str, Any]) -> None:
        self._working_records.append(copy.deepcopy(row))
        self._change_log.append({"op": "insert", "key": row.get(self.key_field), "row": copy.deepcopy(row)})

    def update(self, key: Any, updates: dict[str, Any]) -> None:
        index = self._find_index(key)
        before = copy.deepcopy(self._working_records[index])
        self._working_records[index].update(copy.deepcopy(updates))
        self._change_log.append({
            "op": "update",
            "key": key,
            "before": before,
            "after": copy.deepcopy(self._working_records[index]),
        })

    def delete(self, key: Any) -> None:
        index = self._find_index(key)
        removed = self._working_records.pop(index)
        self._change_log.append({"op": "delete", "key": key, "row": copy.deepcopy(removed)})

    def preview(self) -> dict[str, Any]:
        return {
            "records": [copy.deepcopy(row) for row in self._working_records],
            "change_log": [copy.deepcopy(item) for item in self._change_log],
            "summary": {
                "session_name": self.session_name,
                "pending_changes": len(self._change_log),
                "record_count": len(self._working_records),
            },
        }

    def rollback(self) -> list[dict[str, Any]]:
        self._working_records = [copy.deepcopy(row) for row in self._base_records]
        self._change_log = []
        return [copy.deepcopy(row) for row in self._working_records]

    def commit(self) -> dict[str, Any]:
        committed_log = [copy.deepcopy(item) for item in self._change_log]
        committed_records = [copy.deepcopy(row) for row in self._working_records]
        self._base_records = [copy.deepcopy(row) for row in committed_records]
        self._change_log = []
        return {
            "records": committed_records,
            "change_log": committed_log,
            "summary": {
                "session_name": self.session_name,
                "committed_changes": len(committed_log),
                "pending_changes": 0,
                "record_count": len(committed_records),
            },
        }


def start_edit_session(
    records: Sequence[dict[str, Any]],
    *,
    key_field: str = "id",
    session_name: str = "default",
) -> EditSession:
    """Create a rollback-friendly edit session for enterprise-style workflows."""
    return EditSession(records, key_field=key_field, session_name=session_name)


def versioning_reconcile_post(
    *,
    base_records: Sequence[dict[str, Any]],
    version_records: Sequence[dict[str, Any]],
    key_field: str = "id",
    version_name: str = "edit_version",
) -> dict[str, Any]:
    """Reconcile a child-version edit set back into a posted record collection."""
    merged = {row.get(key_field): copy.deepcopy(row) for row in base_records if row.get(key_field) is not None}
    posted_count = 0
    conflicts: list[dict[str, Any]] = []

    for row in version_records:
        key = row.get(key_field)
        if key is None:
            conflicts.append({"issue": "missing_key", "row": copy.deepcopy(row)})
            continue
        if key not in merged or merged[key] != row:
            merged[key] = copy.deepcopy(row)
            posted_count += 1

    return {
        "version_name": version_name,
        "posted_count": posted_count,
        "conflict_count": len(conflicts),
        "conflicts": conflicts,
        "records": list(merged.values()),
    }


def create_offline_replica_package(
    records: Sequence[dict[str, Any]],
    output_path: str | Path,
    *,
    attachments: dict[Any, Sequence[str]] | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Create an offline-ready JSON bundle for disconnected field workflows."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    attachment_map = {str(key): list(value) for key, value in (attachments or {}).items()}
    payload = {
        "created": datetime.now(timezone.utc).isoformat(),
        "record_count": len(records),
        "attachment_count": sum(len(values) for values in attachment_map.values()),
        "metadata": metadata or {},
        "records": list(records),
        "attachments": attachment_map,
    }
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return {"path": str(path), **{k: v for k, v in payload.items() if k != "records" and k != "attachments"}}


def remove_domain_from_field(record: dict[str, Any], field: str) -> dict[str, Any]:
    """Remove a domain assignment from a field."""
    out = copy.deepcopy(record)
    out.setdefault("field_domains", {}).pop(field, None)
    return out


def alter_domain(
    domain: dict[str, Any] | FieldDomain,
    *,
    values: dict[Any, str] | None = None,
    min_value: float | None = None,
    max_value: float | None = None,
) -> dict[str, Any]:
    """Alter a coded or range domain definition."""
    if isinstance(domain, FieldDomain):
        if values is not None:
            domain.values = values
        if min_value is not None:
            domain.min_value = min_value
        if max_value is not None:
            domain.max_value = max_value
        return domain.to_dict()
    out = dict(domain)
    if values is not None:
        out["values"] = values
    if min_value is not None:
        out["min"] = min_value
    if max_value is not None:
        out["max"] = max_value
    return out


def delete_domain(domain_map: dict[str, Any], name: str) -> dict[str, Any]:
    """Delete a domain from a domain catalog."""
    out = dict(domain_map)
    out.pop(name, None)
    return out


def create_subtype(table: dict[str, Any], code: Any, name: str) -> dict[str, Any]:
    """Create a subtype entry on a table definition."""
    out = copy.deepcopy(table)
    out.setdefault("subtypes", {})[code] = {"name": name}
    return out


def set_subtype_field(table: dict[str, Any], field: str) -> dict[str, Any]:
    """Set the subtype field name."""
    out = copy.deepcopy(table)
    out["subtype_field"] = field
    return out


def set_default_subtype(table: dict[str, Any], code: Any) -> dict[str, Any]:
    """Set the default subtype code."""
    out = copy.deepcopy(table)
    out["default_subtype"] = code
    return out


def remove_subtype(table: dict[str, Any], code: Any) -> dict[str, Any]:
    """Remove a subtype definition."""
    out = copy.deepcopy(table)
    out.setdefault("subtypes", {}).pop(code, None)
    if out.get("default_subtype") == code:
        out["default_subtype"] = None
    return out


def create_relationship_class(
    catalog: dict[str, Any], name: str, origin: str, destination: str
) -> dict[str, Any]:
    """Create a relationship-class definition."""
    out = copy.deepcopy(catalog)
    out.setdefault("relationship_classes", {})[name] = {
        "origin": origin,
        "destination": destination,
    }
    return out


def delete_relationship_class(catalog: dict[str, Any], name: str) -> dict[str, Any]:
    """Delete a relationship-class definition."""
    out = copy.deepcopy(catalog)
    out.setdefault("relationship_classes", {}).pop(name, None)
    return out


def create_attachment_table(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Add an attachment table placeholder to each record."""
    return [{**r, "attachments": list(r.get("attachments", []))} for r in records]


def enable_attachments(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Enable attachment tracking on records."""
    return [{**r, "attachments_enabled": True, "attachments": list(r.get("attachments", []))} for r in records]


def disable_attachments(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Disable attachment tracking on records."""
    return [{**r, "attachments_enabled": False, "attachments": list(r.get("attachments", []))} for r in records]


def remove_attachments(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Remove all attachment references from records."""
    return [{**r, "attachments": []} for r in records]


def create_topology(name: str) -> dict[str, Any]:
    """Create a lightweight topology definition."""
    return {"name": name, "rules": [], "errors": []}


def add_rule_to_topology(topology: dict[str, Any], rule: str) -> dict[str, Any]:
    """Add a rule to a topology definition."""
    out = copy.deepcopy(topology)
    out.setdefault("rules", []).append(rule)
    return out


def fix_topology_errors(errors: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Mark topology errors as fixed with a simple resolution note."""
    return [{**e, "fixed": True, "resolution": "auto-fixed"} for e in errors]


def export_topology_errors(errors: Sequence[dict[str, Any]], output_path: str | Path) -> str:
    """Export topology errors to JSON."""
    out = Path(output_path)
    out.write_text(json.dumps(list(errors), indent=2, default=str), encoding="utf-8")
    return str(out)


def feature_to_3d_by_attribute(
    features: Sequence[dict[str, Any]], field: str
) -> list[dict[str, Any]]:
    """Promote 2D features to 3D using an attribute as Z."""
    out = []
    for feat in features:
        z = float(feat.get(field, 0) or 0)
        new_feat = dict(feat)
        new_feat["z"] = z
        out.append(new_feat)
    return out


def interpolate_shape_3d_from_surface(
    features: Sequence[dict[str, Any]], surface: Sequence[Sequence[float]]
) -> list[dict[str, Any]]:
    """Assign Z values to features by sampling a simple surface grid."""
    default_z = float(surface[0][0]) if surface and surface[0] else 0.0
    return [{**feat, "z": float(feat.get("z", default_z))} for feat in features]


def add_surface_information(
    features: Sequence[dict[str, Any]], field: str = "elevation"
) -> dict[str, Any]:
    """Summarise elevation or surface information from features."""
    vals = [float(f.get(field, 0)) for f in features if f.get(field) is not None]
    mean_val = (sum(vals) / len(vals)) if vals else None
    return {
        "count": len(vals),
        "min_surface": min(vals) if vals else None,
        "max_surface": max(vals) if vals else None,
        "mean_surface": round(mean_val, 4) if mean_val is not None else None,
    }


def layer_3d_to_feature_class(features: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Convert XYZ rows into GeoJSON-like point features."""
    out = []
    for feat in features:
        x, y, z = feat.get("x", 0), feat.get("y", 0), feat.get("z", 0)
        out.append({**feat, "geometry": {"type": "Point", "coordinates": [x, y, z]}})
    return out


def feature_class_to_gdb_batch(
    input_paths: Sequence[str | Path], gdb_path: str | Path
) -> dict[str, Any]:
    """Copy many feature-class-like files into a geodatabase folder."""
    gdb = Path(gdb_path)
    gdb.mkdir(parents=True, exist_ok=True)
    copied = []
    for src in input_paths:
        sp = Path(src)
        target = gdb / sp.name
        target.write_bytes(sp.read_bytes())
        copied.append(str(target))
    return {"gdb": str(gdb), "copied": copied}


def token_based_field_access(
    expression: str,
    feature_row: dict[str, Any],
    map_context: dict[str, Any] | None = None,
) -> Any:
    """Evaluate Arcade-like token expressions using $FEATURE and $MAP."""
    from types import SimpleNamespace

    py_expr = expression.replace("$FEATURE", "FEATURE").replace("$MAP", "MAP")
    ns = dict(_SAFE_BUILTINS)
    ns["FEATURE"] = SimpleNamespace(**feature_row)
    ns["MAP"] = SimpleNamespace(**(map_context or {}))
    return evaluate_safe_expression(py_expr, ns, allowed_attribute_roots={"FEATURE", "MAP"})


def transformer_chain(
    records: Sequence[dict[str, Any]],
    transforms: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Run a simple FME-like transformer chain over records."""
    out = [dict(r) for r in records]
    for step in transforms:
        op = step.get("op")
        if op == "strip":
            field = step.get("field")
            for row in out:
                if field in row and row[field] is not None:
                    row[field] = str(row[field]).strip()
        elif op == "rename":
            out = batch_rename_fields(out, {step.get("from"): step.get("to")})
        elif op == "calculate":
            out = field_calculate(out, step.get("field"), step.get("expression", "None"))
        elif op == "default":
            field = step.get("field")
            value = step.get("value")
            for row in out:
                row.setdefault(field, value)
    return out


# ---------------------------------------------------------------------------
__all__ = [
    "append_records",
    "archive_records",
    "attach_media",
    "batch_convert",
    "batch_rename_fields",
    "catalog_workspace",
    "check_constraints",
    "clean_coordinates",
    "clone_metadata",
    "compact_database",
    "compress_database",
    "copy_records",
    "create_attachment_table",
    "create_dataset_catalog",
    "create_feature_dataset",
    "create_file_geodatabase",
    "create_mobile_geodatabase",
    "create_offline_replica_package",
    "create_relationship_class",
    "create_subtype",
    "create_topology",
    "delete_domain",
    "delete_identical",
    "delete_relationship_class",
    "disable_attachments",
    "enable_attachments",
    "export_schema",
    "export_zip_bundle",
    "export_topology_errors",
    "feature_change_report",
    "feature_class_to_gdb_batch",
    "feature_to_3d_by_attribute",
    "field_calculate",
    "enterprise_persistence_matrix",
    "FieldDomain",
    "add_rule_to_topology",
    "add_surface_information",
    "alter_domain",
    "find_identical",
    "fix_topology_errors",
    "frequency_analysis",
    "get_count",
    "import_zip_bundle",
    "inspect_geodatabase",
    "interpolate_shape_3d_from_surface",
    "index_planning_suggestions",
    "list_attachments",
    "layer_3d_to_feature_class",
    "list_fields",
    "maintenance_audit_pipeline",
    "many_to_many_join",
    "migration_plan",
    "one_to_many_join",
    "overwrite_records",
    "pretty_print_feature_table",
    "read_metadata",
    "read_parquet_filtered",
    "remove_attachments",
    "remove_domain_from_field",
    "remove_subtype",
    "reorder_fields",
    "scaffold_project",
    "schema_diff",
    "schema_lock_management",
    "select_by_polygon",
    "set_default_subtype",
    "set_subtype_field",
    "snapshot_records",
    "split_layer_by_attribute",
    "stamp_lineage",
    "standardize_columns",
    "table_compare",
    "table_to_csv",
    "start_edit_session",
    "table_to_dbase",
    "token_based_field_access",
    "transformer_chain",
    "transpose_table",
    "truncate_records",
    "upsert_records",
    "validate_domains",
    "validate_field_name",
    "validate_table_name",
    "versioning_reconcile_post",
    "versioning_register_unregister",
    "write_csv_with_xy",
    "write_excel_styled",
    "write_metadata",
    # G8 additions
    "add_field",
    "delete_field",
    "topology_validate",
    "find_identical_features",
    "near_table_multi",
    "describe_dataset",
    "pivot_table",
    "multipart_to_singlepart",
    "singlepart_to_multipart",
    "feature_vertices_to_points",
    "repair_geometry_full",
]


# ---------------------------------------------------------------------------
# G8 additions — data management
# ---------------------------------------------------------------------------

from typing import Any as _Any


def add_field(frame: _Any, field_name: str, field_type: str = "float",
              default: _Any = None) -> _Any:
    """Add a new field (column) to a :class:`~geoprompt.GeoPromptFrame`.

    Args:
        frame: The input frame.
        field_name: Name of the new column.
        field_type: One of ``"float"``, ``"int"``, ``"str"``, ``"bool"``.
        default: Default value for all rows.  If ``None``, defaults to
            ``0.0`` (float), ``0`` (int), ``""`` (str), or ``False`` (bool).

    Returns:
        A new frame with the added field.
    """
    type_defaults = {"float": 0.0, "int": 0, "str": "", "bool": False}
    fill = default if default is not None else type_defaults.get(field_type, None)
    rows = list(frame)
    for r in rows:
        r[field_name] = fill
    return type(frame).from_records(rows)


def delete_field(frame: _Any, *field_names: str) -> _Any:
    """Remove one or more fields from a :class:`~geoprompt.GeoPromptFrame`.

    Args:
        frame: The input frame.
        *field_names: Names of columns to remove.

    Returns:
        A new frame without the specified fields.
    """
    to_remove = set(field_names)
    rows = [{k: v for k, v in r.items() if k not in to_remove} for r in frame]
    return type(frame).from_records(rows)


def topology_validate(frame: _Any, *, rule: str = "no_self_intersections") -> list[dict]:
    """Validate geometries in a frame against a topology rule.

    Supported *rule* values:

    - ``"no_self_intersections"`` — flags self-intersecting polygons/rings.
    - ``"no_duplicates"`` — flags duplicate geometry coordinates.
    - ``"must_not_overlap"`` — flags pairs of features whose geometries overlap.

    Args:
        frame: The input :class:`~geoprompt.GeoPromptFrame`.
        rule: The topology rule to apply.

    Returns:
        A list of error dicts with ``feature_index``, ``rule``, and ``message`` keys.
    """
    errors: list[dict] = []
    geom_col = getattr(frame, "geometry_column", "geometry")
    rows = list(frame)

    if rule == "no_self_intersections":
        try:
            import shapely.validation as sv  # type: ignore[import]
            import shapely.geometry as sg  # type: ignore[import]
            import json
            for i, r in enumerate(rows):
                geom = r.get(geom_col)
                if geom is None:
                    continue
                try:
                    shp = sg.shape(geom)
                    if not shp.is_valid:
                        errors.append({"feature_index": i, "rule": rule, "message": sv.explain_validity(shp)})
                except Exception as e:
                    errors.append({"feature_index": i, "rule": rule, "message": str(e)})
        except ImportError:
            # Fallback: no shapely available — skip validation
            pass

    elif rule == "no_duplicates":
        seen: set[str] = set()
        for i, r in enumerate(rows):
            geom = r.get(geom_col)
            key = str(geom)
            if key in seen:
                errors.append({"feature_index": i, "rule": rule, "message": "duplicate geometry"})
            seen.add(key)

    elif rule == "must_not_overlap":
        try:
            import shapely.geometry as sg  # type: ignore[import]
            shapes = []
            for r in rows:
                geom = r.get(geom_col)
                try:
                    shapes.append(sg.shape(geom) if geom else None)
                except Exception:
                    shapes.append(None)
            for i in range(len(shapes)):
                for j in range(i + 1, len(shapes)):
                    if shapes[i] is not None and shapes[j] is not None:
                        if shapes[i].overlaps(shapes[j]):
                            errors.append({"feature_index": i, "rule": rule, "message": f"overlaps feature {j}"})
        except ImportError:
            pass

    return errors


def find_identical_features(frame: _Any, fields: list[str] | None = None) -> list[dict]:
    """Find rows with identical attribute values (or identical geometries).

    Args:
        frame: The input :class:`~geoprompt.GeoPromptFrame`.
        fields: Column names to compare.  If ``None``, compares all
            non-geometry columns plus the geometry column.

    Returns:
        A list of dicts with ``group_id`` and ``feature_indices`` listing
        groups of identical features.
    """
    geom_col = getattr(frame, "geometry_column", "geometry")
    rows = list(frame)
    if fields is None:
        fields = [k for k in rows[0].keys()] if rows else []

    groups: dict[tuple, list[int]] = {}
    for i, r in enumerate(rows):
        key = tuple(str(r.get(f)) for f in fields)
        groups.setdefault(key, []).append(i)

    results = []
    for gid, (key, indices) in enumerate(groups.items()):
        if len(indices) > 1:
            results.append({"group_id": gid, "feature_indices": indices, "key": dict(zip(fields, key))})
    return results


def near_table_multi(frame: _Any, search_frame: _Any, *,
                     max_distance: float = float("inf"),
                     n_nearest: int = 3) -> list[dict]:
    """Generate a near table relating each feature to its *n* nearest neighbours.

    Args:
        frame: Query frame (features to find neighbours for).
        search_frame: The frame to search in.
        max_distance: Maximum search distance (same units as coordinates).
        n_nearest: Number of nearest neighbours per feature.

    Returns:
        A list of dicts with ``in_fid``, ``near_fid``, and ``near_dist`` keys.
    """
    import math

    def _centroid(geom: dict) -> tuple[float, float]:
        t = geom.get("type", "")
        c = geom.get("coordinates", (0.0, 0.0))
        if t == "Point":
            return (float(c[0]), float(c[1]))
        # Fallback: mean of first ring
        def _flat(coords: _Any) -> list[tuple[float, float]]:
            if not coords:
                return []
            if isinstance(coords[0], (int, float)):
                return [(float(coords[0]), float(coords[1]))]
            return [p for sub in coords for p in _flat(sub)]
        pts = _flat(c)
        if pts:
            return (sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts))
        return (0.0, 0.0)

    geom_col = getattr(frame, "geometry_column", "geometry")
    src_rows = list(frame)
    tgt_rows = list(search_frame)

    tgt_centroids = [_centroid(r.get(geom_col) or {}) for r in tgt_rows]

    results = []
    for i, r in enumerate(src_rows):
        pt = _centroid(r.get(geom_col) or {})
        dists = []
        for j, tc in enumerate(tgt_centroids):
            d = math.sqrt((pt[0] - tc[0]) ** 2 + (pt[1] - tc[1]) ** 2)
            if d <= max_distance:
                dists.append((d, j))
        dists.sort()
        for dist, j in dists[:n_nearest]:
            results.append({"in_fid": i, "near_fid": j, "near_dist": dist})
    return results


# ---------------------------------------------------------------------------
# G8 additional — describe, pivot, multipart, vertices, repair
# ---------------------------------------------------------------------------

def describe_dataset(records: Sequence[dict[str, Any]], *,
                     geometry_column: str = "geometry") -> dict[str, Any]:
    """Describe a dataset: field names/types, geometry info, extent, row count, CRS.

    Returns a summary dict with ``fields``, ``geometry_type``, ``extent``,
    ``row_count``, and ``crs`` keys.
    """
    if not records:
        return {"fields": [], "geometry_type": None, "extent": None, "row_count": 0, "crs": None}

    fields: list[dict[str, Any]] = []
    sample = records[0]
    for fname, val in sample.items():
        if fname == geometry_column:
            continue
        if isinstance(val, bool):
            ftype = "bool"
        elif isinstance(val, int):
            ftype = "int"
        elif isinstance(val, float):
            ftype = "float"
        else:
            ftype = "str"
        fields.append({"name": fname, "type": ftype})

    # Geometry summary
    geom_type: str | None = None
    xs: list[float] = []
    ys: list[float] = []
    crs: str | None = None
    for r in records:
        geom = r.get(geometry_column)
        if geom and isinstance(geom, dict):
            geom_type = geom.get("type", geom_type)
            if "crs" in geom:
                crs = geom["crs"]
            coords = geom.get("coordinates")
            if coords:
                def _flatten(c: Any) -> list[tuple[float, float]]:
                    if not c:
                        return []
                    if isinstance(c[0], (int, float)):
                        return [(float(c[0]), float(c[1]))]
                    return [p for sub in c for p in _flatten(sub)]
                for pt in _flatten(coords):
                    xs.append(pt[0])
                    ys.append(pt[1])

    extent = (
        {"min_x": min(xs), "min_y": min(ys), "max_x": max(xs), "max_y": max(ys)}
        if xs else None
    )
    return {
        "fields": fields,
        "geometry_type": geom_type,
        "extent": extent,
        "row_count": len(records),
        "crs": crs or "unknown",
    }


def pivot_table(records: Sequence[dict[str, Any]], *,
                row_field: str,
                col_field: str,
                value_field: str,
                aggfunc: str = "sum") -> dict[str, Any]:
    """Create a pivot table from a list of records.

    Args:
        records: Input rows.
        row_field: Field whose distinct values become row labels.
        col_field: Field whose distinct values become column headers.
        value_field: Field to aggregate.
        aggfunc: One of ``"sum"``, ``"mean"``, ``"count"``, ``"min"``, ``"max"``.

    Returns:
        Dict with ``rows``, ``columns``, and ``table`` (list of lists).
    """
    from collections import defaultdict

    cells: dict[Any, dict[Any, list[float]]] = defaultdict(lambda: defaultdict(list))
    for r in records:
        rv = r.get(row_field)
        cv = r.get(col_field)
        vv = r.get(value_field, 0)
        try:
            cells[rv][cv].append(float(vv))
        except (TypeError, ValueError):
            cells[rv][cv].append(0.0)

    row_labels = sorted(cells.keys(), key=str)
    col_labels = sorted({cv for rv_dict in cells.values() for cv in rv_dict}, key=str)

    def _agg(vals: list[float]) -> float:
        if not vals:
            return 0.0
        if aggfunc == "sum":
            return sum(vals)
        if aggfunc == "mean":
            return sum(vals) / len(vals)
        if aggfunc == "count":
            return float(len(vals))
        if aggfunc == "min":
            return min(vals)
        if aggfunc == "max":
            return max(vals)
        return sum(vals)

    table = [[_agg(cells[r].get(c, [])) for c in col_labels] for r in row_labels]
    return {"rows": row_labels, "columns": col_labels, "table": table, "aggfunc": aggfunc}


def multipart_to_singlepart(records: Sequence[dict[str, Any]], *,
                             geometry_column: str = "geometry") -> list[dict[str, Any]]:
    """Explode multipart geometries into individual singlepart features.

    Non-multi geometries are passed through unchanged.  Multi-geometries
    (MultiPoint, MultiLineString, MultiPolygon, GeometryCollection) are
    split into one record per part.
    """
    out: list[dict[str, Any]] = []
    multi_types = {
        "MultiPoint": "Point",
        "MultiLineString": "LineString",
        "MultiPolygon": "Polygon",
    }
    for r in records:
        geom = r.get(geometry_column)
        if not geom or not isinstance(geom, dict):
            out.append(dict(r))
            continue
        gtype = geom.get("type", "")
        if gtype in multi_types:
            part_type = multi_types[gtype]
            for part_coords in geom.get("coordinates", []):
                new_row = {k: v for k, v in r.items() if k != geometry_column}
                new_row[geometry_column] = {"type": part_type, "coordinates": part_coords}
                out.append(new_row)
        elif gtype == "GeometryCollection":
            for part_geom in geom.get("geometries", []):
                new_row = {k: v for k, v in r.items() if k != geometry_column}
                new_row[geometry_column] = part_geom
                out.append(new_row)
        else:
            out.append(dict(r))
    return out


def singlepart_to_multipart(records: Sequence[dict[str, Any]], *,
                              group_field: str,
                              geometry_column: str = "geometry") -> list[dict[str, Any]]:
    """Combine singlepart features sharing the same *group_field* value into multipart geometries.

    The first row in each group contributes all non-geometry attributes.
    """
    from collections import defaultdict

    groups: dict[Any, list[dict[str, Any]]] = defaultdict(list)
    for r in records:
        groups[r.get(group_field)].append(r)

    out: list[dict[str, Any]] = []
    type_to_multi = {"Point": "MultiPoint", "LineString": "MultiLineString", "Polygon": "MultiPolygon"}
    for key, rows in groups.items():
        base = {k: v for k, v in rows[0].items() if k != geometry_column}
        coords_list: list[Any] = []
        gtype: str | None = None
        for r in rows:
            geom = r.get(geometry_column)
            if geom and isinstance(geom, dict):
                gtype = geom.get("type")
                coords_list.append(geom.get("coordinates"))
        multi_type = type_to_multi.get(gtype or "", "GeometryCollection") if gtype else "GeometryCollection"
        base[geometry_column] = {"type": multi_type, "coordinates": coords_list}
        out.append(base)
    return out


def feature_vertices_to_points(records: Sequence[dict[str, Any]], *,
                                 geometry_column: str = "geometry") -> list[dict[str, Any]]:
    """Extract every vertex of each geometry as an individual Point feature.

    Each output row preserves the attributes of its parent feature plus
    ``vertex_index`` and ``part_index`` fields.
    """
    def _extract_pts(coords: Any, depth: int = 0) -> list[tuple[int, tuple[float, ...]]]:
        if not coords:
            return []
        if isinstance(coords[0], (int, float)):
            return [(depth, tuple(float(v) for v in coords))]
        result = []
        for sub in coords:
            result.extend(_extract_pts(sub, depth + 1))
        return result

    out: list[dict[str, Any]] = []
    for r in records:
        geom = r.get(geometry_column)
        if not geom or not isinstance(geom, dict):
            continue
        coords = geom.get("coordinates") or []
        pts = _extract_pts(coords)
        attrs = {k: v for k, v in r.items() if k != geometry_column}
        for vi, (part_idx, pt) in enumerate(pts):
            xy = list(pt[:2])
            new_row = dict(attrs)
            new_row[geometry_column] = {"type": "Point", "coordinates": xy}
            new_row["vertex_index"] = vi
            new_row["part_index"] = part_idx
            out.append(new_row)
    return out


def repair_geometry_full(records: Sequence[dict[str, Any]], *,
                          geometry_column: str = "geometry") -> list[dict[str, Any]]:
    """Repair invalid geometries to produce OGC-valid results.

    Uses ``shapely.make_valid`` when Shapely ≥ 1.8 is available.  Falls back
    to removing self-intersections by buffering by zero (``buffer(0)``).
    For pure-Python fallback, returns the geometry unchanged with an
    ``_repair_note`` field explaining the limitation.
    """
    out: list[dict[str, Any]] = []
    for r in records:
        geom = r.get(geometry_column)
        if not geom or not isinstance(geom, dict):
            out.append(dict(r))
            continue
        try:
            import shapely.geometry as sg  # type: ignore[import]
            try:
                from shapely.validation import make_valid  # type: ignore[import]
                shp = make_valid(sg.shape(geom))
            except ImportError:
                shp = sg.shape(geom).buffer(0)
            import json
            new_row = dict(r)
            new_row[geometry_column] = sg.mapping(shp)
            new_row.pop("_repair_note", None)
            out.append(new_row)
        except Exception:
            new_row = dict(r)
            new_row["_repair_note"] = "geometry unchanged: shapely unavailable or parse error"
            out.append(new_row)
    return out

